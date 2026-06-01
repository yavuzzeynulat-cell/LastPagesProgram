"""
LastPagesApp - Cube block creator GUI (self-contained).

Tek dosya: writer + GUI + blok verisi.
PyInstaller ile .exe yapilir, Python kurmaya gerek yok.
"""
import json
import os
import re
import shutil
import subprocess
import sys
import tempfile
import threading
import traceback
import urllib.request
from copy import copy
from datetime import date, datetime, timedelta

import tkinter as tk
from tkinter import filedialog, scrolledtext, ttk, messagebox

import openpyxl
from openpyxl.styles import Alignment, Font


__version__ = "0.2.7"
GITHUB_REPO = "yavuzzeynulat-cell/LastPagesProgram"
RELEASES_API = f"https://api.github.com/repos/{GITHUB_REPO}/releases/latest"
UPDATE_ASSET_NAME = "LastPagesApp.exe"


# ---------------- BLOCK DATA ----------------
# Runtime'da Gemini'den dolar (PDF SEC + GEMINI butonu).
# run_writer once Excel'de mevcut cube_no'lari okuyup BLOCKS'tan filtreliyor
# (dedup) - Excel'de olan cube'lar atlanir.

BLOCKS = []


# ---------------- WRITER ----------------

COL = {
    'A': 1, 'B': 2, 'C': 3, 'D': 4, 'E': 5, 'F': 6, 'G': 7, 'H': 8,
    'I': 9, 'J': 10, 'K': 11, 'L': 12, 'M': 13, 'N': 14, 'O': 15,
    'P': 16, 'Q': 17,
}
WHOLE_BLOCK_MERGES = ['A', 'B', 'E', 'F', 'H', 'J', 'P']  # Q is per-row (site rows write 'Site' there)
DATE_COLS = ['I', 'K']
DATE_FORMAT = 'DD/MM/YYYY'


def parse_date(s):
    if isinstance(s, datetime):
        return s.date()
    if isinstance(s, date):
        return s
    s = str(s).replace('/', '.').replace('-', '.')
    parts = s.split('.')
    d, m, y = int(parts[0]), int(parts[1]), int(parts[2])
    if y < 100:
        y += 2000
    return date(y, m, d)


def compute_testing_date(sampling, row_type, custom):
    if custom:
        return parse_date(custom)
    days = {'7d': 7, 'cmd': 7, '28d': 28, 'wp': 28, 'ft': 28,
            '1day': 1, '2day': 2}.get(row_type)
    return sampling + timedelta(days=days) if days else None


def find_last_data_row(ws):
    for r in range(ws.max_row, 0, -1):
        for col in (1, 2, 3):
            if ws.cell(row=r, column=col).value not in (None, ''):
                return r
    return 0


def get_existing_cube_nos(ws):
    """Return set of all cube_no's already present in column A.
    Used to skip blocks the user has already entered (dedup)."""
    nos = set()
    for r in range(1, ws.max_row + 1):
        v = ws.cell(row=r, column=1).value
        if isinstance(v, int):
            nos.add(v)
        elif isinstance(v, float) and v == int(v):
            nos.add(int(v))
        elif isinstance(v, str):
            s = v.strip()
            if s.isdigit():
                nos.add(int(s))
    return nos


def find_donor_block(ws, before_row):
    end = before_row - 1
    while end > 0:
        for col in (1, 2, 3):
            if ws.cell(row=end, column=col).value not in (None, ''):
                break
        else:
            end -= 1
            continue
        break
    if end <= 0:
        return None
    start = end
    for r in range(end, 0, -1):
        if ws.cell(row=r, column=1).value not in (None, ''):
            start = r
            for rng in ws.merged_cells.ranges:
                if rng.min_col == 1 and rng.max_col == 1 and rng.min_row == r:
                    end = max(end, rng.max_row)
                    break
            break
    return (start, end)


def get_donor_styles(ws, start, end):
    n = end - start + 1
    styles = {}
    for col_letter, col_idx in COL.items():
        top = ws.cell(row=start, column=col_idx)
        mid = ws.cell(row=start + n // 2, column=col_idx)
        bot = ws.cell(row=end, column=col_idx)
        styles[col_letter] = {'top': top, 'mid': mid, 'bot': bot}
    return styles


def get_donor_formula(ws, start, end, col_letter):
    col_idx = COL[col_letter]
    for r in range(start, end + 1):
        val = ws.cell(row=r, column=col_idx).value
        if isinstance(val, str) and val.startswith('='):
            return (val, r)
    return (None, None)


def shift_formula(formula, src_row, dst_row):
    diff = dst_row - src_row
    if diff == 0:
        return formula
    return re.sub(r'(\$?[A-Z]+\$?)(\d+)',
                  lambda m: f"{m.group(1)}{int(m.group(2)) + diff}",
                  formula)


def copy_style(src_cell, dst_cell):
    # Copy the StyleArray INDEX directly instead of duplicating each
    # style component (font/fill/border/alignment). Otherwise openpyxl
    # creates a fresh style entry per cell -> styles.xml explodes ->
    # Excel becomes very slow on big sheets.
    if not src_cell.has_style:
        return
    dst_cell._style = copy(src_cell._style)


def style_position(i, n):
    if i == 0:
        return 'top'
    if i == n - 1:
        return 'bot'
    return 'mid'


def unmerge_overlapping(ws, start_row, end_row):
    to_remove = [str(r) for r in list(ws.merged_cells.ranges)
                 if r.min_row <= end_row and r.max_row >= start_row]
    for r in to_remove:
        ws.unmerge_cells(r)


def write_block(ws, block, start_row, donor_styles, donor_formulas):
    rows = block['rows']
    n = len(rows)
    end_row = start_row + n - 1
    unmerge_overlapping(ws, start_row, end_row)
    sampling = parse_date(block['sampling_date'])
    cmd_idx = next((i for i, r in enumerate(rows) if r.get('row_type') == 'cmd'), None)

    batch_tickets = block.get('batch_tickets')
    if not batch_tickets:
        bt = block.get('batch_ticket')
        batch_tickets = [{'ticket': bt, 'rows': [0, n - 1]}] if bt is not None else []

    for i in range(n):
        r = start_row + i
        pos = style_position(i, n)
        for col_letter, col_idx in COL.items():
            copy_style(donor_styles[col_letter][pos], ws.cell(row=r, column=col_idx))
        for dc in DATE_COLS:
            ws.cell(row=r, column=COL[dc]).number_format = DATE_FORMAT

    if n > 1:
        for col_letter in WHOLE_BLOCK_MERGES:
            ws.merge_cells(start_row=start_row, end_row=end_row,
                           start_column=COL[col_letter], end_column=COL[col_letter])

    # D column: SINGLE merge across whole block with multi-line content
    # ("S2A BP\nCMD-XXXX"). Earlier we split into two merges, but the user
    # confirmed it should be one cell with two lines.
    if n > 1:
        ws.merge_cells(start_row=start_row, end_row=end_row,
                       start_column=COL['D'], end_column=COL['D'])

    for bt in batch_tickets:
        # rows: yarat hem [start,end] hem tam liste [0,1,2,3] gelebilir - min/max al
        r = bt.get('rows') or []
        if not r:
            continue
        a, b = min(r), max(r)
        if b > a:
            ws.merge_cells(start_row=start_row + a, end_row=start_row + b,
                           start_column=COL['G'], end_column=COL['G'])

    ws.cell(row=start_row, column=COL['A'], value=block['cube_no'])
    ws.cell(row=start_row, column=COL['B'], value=block['sample_mark'])
    # D column: single merged cell with multi-line content (supplier + CMD-XXX)
    d_value = block.get('supplier', 'S2A BP')
    if block.get('cmd_code'):
        d_value = f"{d_value}\nCMD-{block['cmd_code']}"
    d_cell = ws.cell(row=start_row, column=COL['D'])
    d_cell.value = d_value
    d_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    e_cell = ws.cell(row=start_row, column=COL['E'], value=block.get('site_location', ''))
    # E column text is RED for newly added blocks so user can spot them at a glance
    src_font = e_cell.font
    e_cell.font = Font(
        name=src_font.name, size=src_font.size,
        bold=src_font.bold, italic=src_font.italic,
        family=src_font.family,
        color='FFFF0000',
    )
    if block.get('section') is not None:
        ws.cell(row=start_row, column=COL['F'], value=block['section'])
    for bt in batch_tickets:
        r = bt.get('rows') or []
        if not r:
            continue
        ws.cell(row=start_row + min(r), column=COL['G'], value=bt.get('ticket'))
    ws.cell(row=start_row, column=COL['H'], value=block.get('c_grade', ''))
    ws.cell(row=start_row, column=COL['J'], value=block.get('sampled_by', ''))

    for i, row in enumerate(rows):
        r = start_row + i
        if row.get('mould') is not None:
            ws.cell(row=r, column=COL['C'], value=row['mould'])
        ws.cell(row=r, column=COL['I'], value=sampling)
        td = compute_testing_date(sampling, row.get('row_type'), row.get('testing_date'))
        if td is not None:
            ws.cell(row=r, column=COL['K'], value=td)
        # Age (L): formula =K-I when both sampling + testing date are real numbers.
        # For WP / F-T / Site without date, write the literal label.
        age_val = row.get('age')
        if td is not None and isinstance(age_val, (int, float)):
            ws.cell(row=r, column=COL['L'], value=f"=K{r}-I{r}")
        elif age_val is not None:
            ws.cell(row=r, column=COL['L'], value=age_val)
        if row.get('row_type') == 'ft' and row.get('ft_note'):
            ws.cell(row=r, column=COL['N'], value=row['ft_note'])
        if row.get('row_type') == 'site':
            ws.cell(row=r, column=COL['Q'], value='Site')
        f, src_r = donor_formulas.get('O', (None, None))
        if f:
            ws.cell(row=r, column=COL['O'], value=shift_formula(f, src_r, r))

    return end_row + 1


def _excel_recalc_and_save(path, log):
    """Open the file in Excel via COM, force full recalc, save, close.
    This restores Excel's formula calculation cache (openpyxl strips it),
    making the file open at original speed instead of recalculating
    25k+ rows on every open."""
    try:
        import pythoncom
        from win32com.client import DispatchEx
    except ImportError:
        log("Uyari: pywin32 yok, recalc adimi atlandi (dosya yine yazildi).")
        return
    pythoncom.CoInitialize()
    excel = None
    try:
        excel = DispatchEx('Excel.Application')
        excel.Visible = False
        excel.DisplayAlerts = False
        excel.ScreenUpdating = False
        wb = excel.Workbooks.Open(os.path.abspath(path))
        try:
            wb.Application.CalculateFull()
            wb.Save()
        finally:
            wb.Close(SaveChanges=False)
    finally:
        if excel is not None:
            try:
                excel.Quit()
            except Exception:
                pass
        try:
            pythoncom.CoUninitialize()
        except Exception:
            pass


def run_writer(excel_path, log):
    log(f"Excel aciliyor: {excel_path}")
    wb = openpyxl.load_workbook(excel_path)
    if 'Concrete' not in wb.sheetnames:
        raise RuntimeError(f"'Concrete' sheet bulunamadi. Mevcut: {wb.sheetnames}")
    ws = wb['Concrete']

    # Dedup: hangi cube_no'lar Excel'de zaten var?
    existing = get_existing_cube_nos(ws)
    log(f"Excel'de mevcut: {len(existing)} cube_no.")

    blocks_to_write = [b for b in BLOCKS if b['cube_no'] not in existing]
    skipped = [b['cube_no'] for b in BLOCKS if b['cube_no'] in existing]
    if skipped:
        log(f"Zaten Excel'de oldugu icin ATLANAN: {skipped}")
    if not blocks_to_write:
        log("Yazilacak yeni blok yok. Excel zaten guncel.")
        return
    log(f"YAZILACAK ({len(blocks_to_write)} blok): {[b['cube_no'] for b in blocks_to_write]}")

    start_row = find_last_data_row(ws) + 1
    log(f"Yazma baslangici: satir {start_row}")

    donor = find_donor_block(ws, start_row)
    if not donor:
        raise RuntimeError("Donor block bulunamadi - Concrete sheet bos mu?")
    donor_start, donor_end = donor
    log(f"Donor block: satir {donor_start}-{donor_end}")

    donor_styles = get_donor_styles(ws, donor_start, donor_end)
    donor_formulas = {'O': get_donor_formula(ws, donor_start, donor_end, 'O')}
    if donor_formulas['O'][0]:
        log(f"Donor O formul: {donor_formulas['O'][0]} (@ satir {donor_formulas['O'][1]})")

    next_row = start_row
    for block in blocks_to_write:
        log(f"  -> cube {block['cube_no']} @ satir {next_row} ({len(block['rows'])} satir)")
        next_row = write_block(ws, block, next_row, donor_styles, donor_formulas)

    log(f"Kaydediliyor: {excel_path}")
    wb.save(excel_path)
    log("Excel ile recalc + kaydet (dosyanin hizini geri getirir)...")
    try:
        _excel_recalc_and_save(excel_path, log)
        log("Recalc tamam - dosya artik hizli aciliyor.")
    except Exception as e:
        log(f"Uyari: recalc basarisiz ({e}). Veri yazildi ama ilk acilis biraz yavas olabilir.")
    log("Tamamlandi.")


# ---------------- AUTO-UPDATER ----------------

def _parse_version(s):
    try:
        return tuple(int(p) for p in s.lstrip('v').strip().split('.'))
    except (ValueError, AttributeError):
        return None


def _is_newer(latest, current):
    a, b = _parse_version(latest), _parse_version(current)
    return a is not None and b is not None and a > b


def _get_latest_release():
    """Return (tag_no_v, asset_url) or None on any failure."""
    try:
        req = urllib.request.Request(
            RELEASES_API,
            headers={'User-Agent': f'LastPagesApp/{__version__}'},
        )
        with urllib.request.urlopen(req, timeout=5) as resp:
            data = json.load(resp)
        tag = data.get('tag_name', '').lstrip('v').strip()
        if not tag:
            return None
        for a in data.get('assets', []):
            if a.get('name') == UPDATE_ASSET_NAME:
                return (tag, a.get('browser_download_url'))
        return None
    except Exception:
        return None


def _download_file(url, dest_path, progress_cb=None):
    """Download to dest_path; raise if size mismatch (Content-Length vs written)."""
    req = urllib.request.Request(url, headers={'User-Agent': f'LastPagesApp/{__version__}'})
    with urllib.request.urlopen(req, timeout=60) as resp:
        total = int(resp.headers.get('Content-Length') or 0)
        downloaded = 0
        with open(dest_path, 'wb') as f:
            while True:
                chunk = resp.read(64 * 1024)
                if not chunk:
                    break
                f.write(chunk)
                downloaded += len(chunk)
                if progress_cb:
                    progress_cb(downloaded, total)
    actual = os.path.getsize(dest_path)
    if total and actual != total:
        raise IOError(f"Indirme eksik: {actual} / {total} bayt")
    if actual < 1024 * 1024:  # exe should be >1MB
        raise IOError(f"Indirilen dosya cok kucuk: {actual} bayt (bozuk olabilir)")


PS_UPDATER_TEMPLATE = r'''$ErrorActionPreference = "Stop"
$log = "__LOG__"
function Log($msg) { Add-Content -Path $log -Value "$(Get-Date -Format o) $msg" }
Log "updater started, waiting for pid __PID__"
$waited = 0
while (Get-Process -Id __PID__ -ErrorAction SilentlyContinue) {
    Start-Sleep -Milliseconds 500
    $waited += 0.5
    if ($waited -gt 20) { Log "timeout waiting for app"; break }
}
Start-Sleep -Seconds 2  # let Windows fully release file handles
$newSize = (Get-Item -LiteralPath "__NEW_EXE__").Length
Log "new exe size: $newSize bytes"
$copied = $false
for ($i = 1; $i -le 5; $i++) {
    try {
        Copy-Item -LiteralPath "__NEW_EXE__" -Destination "__TARGET_EXE__" -Force
        $destSize = (Get-Item -LiteralPath "__TARGET_EXE__").Length
        if ($destSize -ne $newSize) {
            throw "size mismatch after copy: $destSize vs $newSize"
        }
        Log "copied (attempt $i): $destSize bytes -> __TARGET_EXE__"
        $copied = $true
        break
    } catch {
        Log "copy attempt $i failed: $_"
        Start-Sleep -Seconds 2
    }
}
if (-not $copied) {
    Log "ABORT: could not replace exe after 5 attempts; restarting OLD version"
    Start-Process -FilePath "__TARGET_EXE__"
    exit 1
}
try {
    Start-Process -FilePath "__TARGET_EXE__"
    Log "restarted"
} catch {
    Log "ERROR starting new exe: $_"
}
'''


def _write_powershell_updater(temp_dir, new_exe, target_exe, log_path):
    ps_path = os.path.join(temp_dir, 'update.ps1')
    script = (PS_UPDATER_TEMPLATE
              .replace('__LOG__', log_path)
              .replace('__PID__', str(os.getpid()))
              .replace('__NEW_EXE__', new_exe)
              .replace('__TARGET_EXE__', target_exe))
    with open(ps_path, 'w', encoding='utf-8') as f:
        f.write(script)
    return ps_path


def _launch_powershell_updater(ps_path):
    # CREATE_NO_WINDOW = 0x08000000 — helper runs silently
    subprocess.Popen(
        ['powershell', '-ExecutionPolicy', 'Bypass',
         '-WindowStyle', 'Hidden', '-File', ps_path],
        creationflags=0x08000000,
    )


def _do_update(root, asset_url, new_tag):
    """Tk thread: download into temp, launch helper, exit app."""
    target_exe = sys.executable if getattr(sys, 'frozen', False) else None
    if not target_exe or not target_exe.lower().endswith('.exe'):
        messagebox.showinfo("Guncelleme",
            "Otomatik guncelleme sadece .exe surumde calisir.\n"
            "Mevcut versiyonla devam ediliyor.")
        return

    temp_dir = os.path.join(tempfile.gettempdir(), 'lastpages_update')
    try:
        if os.path.exists(temp_dir):
            shutil.rmtree(temp_dir, ignore_errors=True)
        os.makedirs(temp_dir, exist_ok=True)
    except Exception as e:
        messagebox.showerror("Guncelleme hatasi", f"Temp klasor olusturulamadi: {e}")
        return

    new_exe = os.path.join(temp_dir, UPDATE_ASSET_NAME)
    log_path = os.path.join(temp_dir, 'update.log')

    win = tk.Toplevel(root)
    win.title("Guncelleniyor")
    win.geometry("420x110")
    win.transient(root)
    win.grab_set()
    win.protocol("WM_DELETE_WINDOW", lambda: None)
    ttk.Label(win, text=f"v{new_tag} indiriliyor...", padding=10).pack()
    bar = ttk.Progressbar(win, length=380, mode='determinate', maximum=100)
    bar.pack(padx=20, pady=5)
    pct = ttk.Label(win, text="0%")
    pct.pack()

    state = {'err': None}

    def progress(d, t):
        def update():
            if t:
                bar['maximum'] = t
                bar['value'] = d
                pct.config(text=f"{int(d * 100 / t)}%  ({d // 1024} / {t // 1024} KB)")
            else:
                pct.config(text=f"{d // 1024} KB")
        root.after(0, update)

    def worker():
        try:
            _download_file(asset_url, new_exe, progress)
        except Exception as e:
            state['err'] = str(e)
        root.after(0, finish)

    def finish():
        try:
            win.destroy()
        except Exception:
            pass
        if state['err']:
            messagebox.showerror("Guncelleme hatasi",
                f"Indirme basarisiz:\n{state['err']}\n\nMevcut versiyonla devam ediliyor.")
            return
        ps_path = _write_powershell_updater(temp_dir, new_exe, target_exe, log_path)
        _launch_powershell_updater(ps_path)
        root.destroy()
        sys.exit(0)

    threading.Thread(target=worker, daemon=True).start()


def check_for_update(root):
    """Background-check GitHub, prompt on Tk thread if newer."""
    def bg():
        result = _get_latest_release()
        if not result:
            return
        latest_tag, asset_url = result
        if not asset_url or not _is_newer(latest_tag, __version__):
            return
        def prompt():
            if messagebox.askyesno("Guncelleme mevcut",
                f"Yeni versiyon v{latest_tag} mevcut.\n"
                f"Mevcut: v{__version__}\n\n"
                "Simdi guncellemek ister misin?"):
                _do_update(root, asset_url, latest_tag)
        root.after(0, prompt)
    threading.Thread(target=bg, daemon=True).start()


# ---------------- GEMINI PDF READER ----------------

APP_DIR = os.path.dirname(os.path.abspath(sys.argv[0]))
CONFIG = os.path.join(APP_DIR, "config.txt")
API_KEY_FILE = os.path.join(APP_DIR, "gemini_api_key.txt")
CACHE_DIR = os.path.join(APP_DIR, "gemini_cache")

GEMINI_MODEL = "gemini-2.5-flash"

GEMINI_PROMPT = """Bu el yazisi sayfa(lar)indaki HER cube blogunu JSON listesi olarak cikar.

Cube bloklari kutucuklardir; her blok bir "Cube No" satiriyla baslar ve bir sonraki Cube No'ya kadar surer. Tum sayfalardaki butun bloklari sirayla isle.

ONEMLI - SAYFALAR ARASI BOLUNMUS BLOKLAR:
- Bir cube blogu iki sayfa arasinda bolunmus olabilir (yarisi 1. sayfanin sonunda, devami 2. sayfanin basinda)
- Bu durumda bloku TEK blok olarak cikar - hem ust hem alt satirlari ayni "rows" listesinde birlestir
- 2. sayfanin basindaki rows'un cube_no'su yoktur (ust kisimda kalmistir) - 1. sayfada gordugun son cube_no ile esle
- ASLA ayri iki blok yapma; ayni cube_no'ya ait satirlar tek blok olmali

ROW TYPES:
- "7d"   : 7-gun testi (Age=7)
- "cmd"  : CMD satiri (Concrete Mix Design; D sutununda "CMD" yazili, genelde 3. satir)
- "28d"  : 28-gun testi (Age=28)
- "wp"   : Waiting Period (Age="WP")
- "1day" : 1-gun testi (Age=1)
- "2day" : 2-gun testi (Age=2)
- "site" : Sahaya gonderilen kupler ("Tested By" sutununda "Site" yazili)
- "ft"   : F/T ozet satiri (Age="F/T", Load sutununda "(X Adet)" yazili)

Eger bir satirda hem age=1 (1.Day) hem "Site" varsa: row_type="site" + age=1 + testing_date (sampling+1).
Mould 60 ornek: row_type="site" + age=2 + testing_date.

JSON formati ZORUNLU:
{
  "blocks": [
    {
      "cube_no": 698,
      "sample_mark": "G26-CON-743",
      "supplier": "S2A BP",
      "cmd_code": "7312",
      "site_location": "B0051 Bridge Km:5+050 SBL Link Slab",
      "section": "2A",
      "batch_ticket": 14958,
      "c_grade": "C30/37",
      "sampling_date": "21.05.26",
      "sampled_by": "Y.A",
      "rows": [
        {"mould": 176, "row_type": "7d", "age": 7},
        {"mould": 9, "row_type": "7d", "age": 7},
        {"mould": 136, "row_type": "cmd", "age": 7},
        {"mould": 78, "row_type": "28d", "age": 28},
        {"mould": 142, "row_type": "28d", "age": 28},
        {"mould": 106, "row_type": "28d", "age": 28},
        {"mould": 166, "row_type": "site", "age": 1, "testing_date": "22.05.26"},
        {"mould": 89, "row_type": "site"},
        {"mould": 8, "row_type": "site"}
      ]
    }
  ]
}

Multi-batch bloklarda (G sutununda farkli ticket no'lar varsa) batch_ticket yerine batch_tickets kullan:
"batch_tickets": [
    {"ticket": 14979, "rows": [0, 5]},
    {"ticket": 14983, "rows": [6, 11]}
]
ONEMLI: "rows" alani TAM OLARAK iki sayi olmali: [ilk_satir_index, son_satir_index].
ASLA [0,1,2,3,4,5] gibi tam liste verme - sadece [first, last] iki eleman.
Satir indexleri blok icinde 0-bazli (blogun ilk satiri = 0).

c_grade formati: "C30/37" gibi (basina C ekle).
sampling_date / testing_date formati: "DD.MM.YY" (orn: "21.05.26").
F/T satirinda: {"mould": null, "row_type": "ft", "age": "F/T", "ft_note": "(15 Adet)"}.

SADECE GECERLI JSON dondur, markdown code fence kullanma, baska metin ekleme.
JSON'i COMPACT yaz (indent yok, satir sonu yok, ekstra bosluk yok) - token tasarrufu icin.
Cevap mutlaka } ile bitmeli (yarida kesilmis JSON olmasin)."""


def _pdf_content_hash(path):
    """SHA256 of file content - kisa hex (16 char) doner. PDF degisirse cache invalid."""
    import hashlib
    h = hashlib.sha256()
    with open(path, 'rb') as f:
        for chunk in iter(lambda: f.read(65536), b''):
            h.update(chunk)
    return h.hexdigest()[:16]


def cache_load(pdf_hash):
    """Var ise blocks list'i dondurur, yoksa None."""
    path = os.path.join(CACHE_DIR, f"{pdf_hash}.json")
    if not os.path.exists(path):
        return None
    try:
        with open(path, encoding='utf-8') as f:
            return json.load(f)
    except Exception:
        return None


def cache_save(pdf_hash, blocks, pdf_path):
    """Cache'e yaz."""
    try:
        os.makedirs(CACHE_DIR, exist_ok=True)
        # Metadata icin yardimci dosya
        data = {
            "pdf_name": os.path.basename(pdf_path),
            "pdf_hash": pdf_hash,
            "saved_at": datetime.now().isoformat(timespec='seconds'),
            "block_count": len(blocks),
            "blocks": blocks,
        }
        path = os.path.join(CACHE_DIR, f"{pdf_hash}.json")
        with open(path, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
    except Exception:
        pass


def load_api_key():
    if os.path.exists(API_KEY_FILE):
        try:
            with open(API_KEY_FILE, encoding='utf-8') as f:
                return f.read().strip()
        except Exception:
            pass
    return ""


def save_api_key(key):
    with open(API_KEY_FILE, 'w', encoding='utf-8') as f:
        f.write(key.strip())


def _normalize_block(b, log):
    """Defensive normalizer: Gemini'nin sasirtici formatlarini standartlastir.
    Yan etki: blok dict'i in-place duzenler ve dondurur (ya da None drop)."""
    if not isinstance(b, dict):
        log(f"  uyari: blok dict degil ({type(b).__name__}), atlandi")
        return None

    # cube_no: int olmali
    try:
        b['cube_no'] = int(b['cube_no'])
    except (KeyError, ValueError, TypeError):
        log(f"  uyari: gecersiz cube_no={b.get('cube_no')!r}, blok atlandi")
        return None

    # rows: list olmali
    rows = b.get('rows')
    if not isinstance(rows, list) or len(rows) == 0:
        log(f"  uyari: cube {b['cube_no']} bos/gecersiz rows, atlandi")
        return None

    # Her row'u temizle
    cleaned_rows = []
    for r in rows:
        if not isinstance(r, dict):
            continue
        # mould: int veya None
        m = r.get('mould')
        if m is None or m == "" or m == "null":
            r['mould'] = None
        else:
            try:
                r['mould'] = int(m)
            except (ValueError, TypeError):
                r['mould'] = None
        # row_type: lower-case
        rt = r.get('row_type')
        if isinstance(rt, str):
            r['row_type'] = rt.lower().strip()
        # age: int / str / left as is
        cleaned_rows.append(r)
    b['rows'] = cleaned_rows
    n = len(cleaned_rows)

    # batch_tickets normalize: [start,end] veya tam liste her ikisini de kabul et
    if b.get('batch_tickets'):
        new_bts = []
        for bt in b['batch_tickets']:
            if not isinstance(bt, dict):
                continue
            r = bt.get('rows') or []
            if isinstance(r, int):
                r = [r]
            if not r:
                continue
            try:
                a, c = int(min(r)), int(max(r))
            except (ValueError, TypeError):
                continue
            # range clip [0, n-1]
            a = max(0, min(a, n - 1))
            c = max(0, min(c, n - 1))
            new_bts.append({'ticket': bt.get('ticket'), 'rows': [a, c]})
        b['batch_tickets'] = new_bts
        if not new_bts and 'batch_ticket' not in b:
            log(f"  uyari: cube {b['cube_no']} batch_tickets bos kaldi")

    # batch_ticket tek: int olmasi guzel ama string de olabilir, dokunma

    # section: int veya str olabilir, dokunma
    # c_grade: basinda C yoksa ekle
    cg = b.get('c_grade')
    if isinstance(cg, str) and cg and not cg.upper().startswith('C'):
        b['c_grade'] = 'C' + cg

    # sample_mark, sampling_date, sampled_by, supplier, cmd_code, site_location: as-is
    return b


def _parse_gemini_json(text, log):
    """Tolerantly parse Gemini response. Returns list of blocks."""
    text = (text or "").strip()
    if not text:
        return []
    # Strip markdown code fences (model sometimes ignores instruction)
    if text.startswith("```"):
        parts = text.split("```")
        if len(parts) >= 2:
            text = parts[1]
            if text.startswith("json"):
                text = text[4:]
            text = text.strip()
    # Best-effort: find first '{'/last '}' to skip any pre/post text
    first = text.find('{')
    last = text.rfind('}')
    candidate = text[first:last + 1] if first >= 0 and last > first else text
    # Truncation detection: response should end with '}' or ']'
    stripped_tail = candidate.rstrip()
    truncated = not stripped_tail.endswith(('}', ']'))
    try:
        data = json.loads(candidate)
    except json.JSONDecodeError as e:
        log(f"Gemini JSON parse hatasi @ {e.pos} (toplam {len(text)} karakter)")
        log("--- ILK 400 KARAKTER ---")
        log(text[:400])
        log("--- SON 400 KARAKTER ---")
        log(text[-400:])
        if truncated:
            raise RuntimeError(
                "Gemini cevabi YARIDA KESILDI. Bu sayfada cok fazla blok/satir var, "
                "max_output_tokens limit asilmis olabilir. "
                f"Cevap '{stripped_tail[-30:]!r}' ile bitiyor (kapatici parantez bekleniyordu)."
            )
        raise RuntimeError(f"JSON parse hatasi: {e}")
    if isinstance(data, list):
        return data
    if isinstance(data, dict):
        return data.get('blocks', []) or []
    return []


def gemini_extract_blocks(pdf_path, api_key, log, progress_cb=None, force=False):
    """Render PDF pages and extract blocks via Gemini.

    CACHE: Once PDF'in SHA256'sini bulup gemini_cache/<hash>.json'a bakar.
    Hit ise Gemini hic cagrilmaz (PARA TASARRUFU). force=True ile bypass.

    Robust strategy:
      1) Sayfa basina ayri istek - kucuk payload, hizli yanit
      2) Her istegin timeout'u 300sn
      3) Hata olursa 3x retry, exponential backoff (1, 2, 4 sn)
      4) Image hala buyukse DPI dusur (200 -> 150 -> 120)
      5) Sayfalar arasi cube_no dedup + normalize
    """
    # CACHE check
    pdf_hash = _pdf_content_hash(pdf_path)
    log(f"PDF hash: {pdf_hash}")
    if not force:
        cached = cache_load(pdf_hash)
        if cached is not None:
            blocks = cached.get('blocks') if isinstance(cached, dict) else cached
            if isinstance(blocks, list) and blocks:
                log(f"*** CACHE HIT *** - {len(blocks)} blok daha onceden okunmustu.")
                log(f"  Gemini cagrilmiyor (PARA TASARRUFU). Cube_no'lar: {[b.get('cube_no') for b in blocks]}")
                log(f"  Yeniden okumak istersen cache dosyasini sil: {CACHE_DIR}\\{pdf_hash}.json")
                return blocks
        log("Cache miss - Gemini cagrilacak.")

    import time
    import threading
    import fitz  # pymupdf
    from PIL import Image
    from io import BytesIO
    import google.generativeai as genai
    from google.api_core import exceptions as gax
    from concurrent.futures import ThreadPoolExecutor, as_completed

    genai.configure(api_key=api_key)
    generation_config = {
        "response_mime_type": "application/json",
        "max_output_tokens": 32768,
        "temperature": 0.0,
    }
    model = genai.GenerativeModel(GEMINI_MODEL, generation_config=generation_config)
    request_options = {"timeout": 300}

    log(f"PDF aciliyor: {pdf_path}")
    doc = fitz.open(pdf_path)
    page_count = len(doc)
    log(f"Sayfa sayisi: {page_count}")
    if page_count == 0:
        raise RuntimeError("PDF bos.")

    # Tum sayfalari onceden render et.
    # ONEMLI: DPI degil PIXEL hedefliyoruz. Gemini'nin vision encoder'i icinde
    # 1568px patch'lere boluyor - 2500px sweet spot (yeterli detay + hizli).
    TARGET_MAX_PX = 2500
    FALLBACK_MAX_PX = 3500

    def render_at_target(page, max_px):
        rect = page.rect
        max_pt = max(rect.width, rect.height)
        if max_pt <= 0:
            target_dpi = 150
        else:
            target_dpi = (max_px * 72.0) / max_pt
        target_dpi = max(100, min(target_dpi, 300))
        pm = page.get_pixmap(dpi=target_dpi)
        return pm.tobytes("png"), pm.width, pm.height, int(round(target_dpi))

    log(f"Sayfalar render ediliyor (hedef max {TARGET_MAX_PX}px)...")
    rendered = {}
    total_kb = 0
    for i in range(page_count):
        page = doc[i]
        png_bytes, w, h, used_dpi = render_at_target(page, TARGET_MAX_PX)
        rendered[i] = (png_bytes, w, h, used_dpi)
        total_kb += len(png_bytes) // 1024
        log(f"  sayfa {i + 1}/{page_count}: {w}x{h} @ dpi={used_dpi}, {len(png_bytes) // 1024} KB")
    log(f"Toplam payload: {total_kb} KB (eskiden ~95 MB idi - cok kucuk)")

    log("")
    log(f"TEK Gemini cagrisi - {page_count} sayfa birlikte (sayfalar arasi bloklari yakalar)")
    log("(toplam ~60-90sn beklenir, timeout 300sn)")

    def call_with_retries(prompt_parts, max_attempts=3):
        last_err = None
        for attempt in range(1, max_attempts + 1):
            try:
                return model.generate_content(prompt_parts, request_options=request_options)
            except (gax.DeadlineExceeded, gax.ResourceExhausted,
                    gax.ServiceUnavailable, gax.InternalServerError) as e:
                last_err = e
                wait = 2 ** (attempt - 1)
                log(f"  {type(e).__name__} deneme {attempt}/{max_attempts}, {wait}s bekle")
                if attempt < max_attempts:
                    time.sleep(wait)
            except gax.InvalidArgument:
                raise
            except Exception as e:
                last_err = e
                if attempt < max_attempts:
                    time.sleep(1)
        raise last_err if last_err else RuntimeError("Bilinmeyen hata")

    def process_page_solo(idx):
        """Fallback: tek sayfayi Gemini'ye yolla."""
        png_bytes = rendered[idx][0]
        img = Image.open(BytesIO(png_bytes)); img.load()
        last_err = None
        for target_px, source in [(TARGET_MAX_PX, png_bytes), (FALLBACK_MAX_PX, None)]:
            try:
                if source is None:
                    fresh, _, _, _ = render_at_target(doc[idx], target_px)
                    img2 = Image.open(BytesIO(fresh)); img2.load()
                else:
                    img2 = img
                resp = call_with_retries([GEMINI_PROMPT, img2])
                blocks = _parse_gemini_json(resp.text, log)
                log(f"  [sayfa {idx + 1}] max_px={target_px} -> {len(blocks)} blok")
                return idx, blocks
            except gax.InvalidArgument as e:
                last_err = e
                continue
            except Exception as e:
                last_err = e
                break
        raise RuntimeError(f"Sayfa {idx + 1}: {last_err}")

    # ASIL CAGRI: Tum sayfalar TEK request (sayfa sinirlarini yakalar)
    if progress_cb:
        progress_cb(0, 1, "gemini")
    img_list = []
    for i in range(page_count):
        img = Image.open(BytesIO(rendered[i][0])); img.load()
        img_list.append(img)

    log("Istek gonderiliyor (tek call - tum sayfalar)...")
    all_blocks = None
    try:
        resp = call_with_retries([GEMINI_PROMPT] + img_list)
        all_blocks = _parse_gemini_json(resp.text, log)
        log(f"Tek-call basarili: {len(all_blocks)} blok")
    except (gax.InvalidArgument, RuntimeError) as e:
        # Cok blok varsa response truncate olabilir, ya da payload buyuk olabilir
        # Fallback: sayfa basina paralel
        log(f"Tek-call basarisiz ({type(e).__name__}: {e})")
        log("FALLBACK: sayfa basina paralel (sayfa sinirlari kac olabilir)")
        all_blocks_by_page = {}
        with ThreadPoolExecutor(max_workers=min(page_count, 4)) as pool:
            future_to_idx = {pool.submit(process_page_solo, i): i for i in range(page_count)}
            done = 0
            for fut in as_completed(future_to_idx):
                idx = future_to_idx[fut]
                try:
                    page_idx, blocks = fut.result()
                    all_blocks_by_page[page_idx] = blocks
                except Exception as ex:
                    log(f"  [sayfa {idx + 1}] FAIL: {ex}")
                    all_blocks_by_page[idx] = None
                done += 1
                if progress_cb:
                    progress_cb(done - 1, page_count, "page")
        failed = [i + 1 for i, b in all_blocks_by_page.items() if b is None]
        if failed:
            raise RuntimeError(f"Sayfa(lar) okunamadi: {failed}")
        all_blocks = []
        for i in range(page_count):
            all_blocks.extend(all_blocks_by_page.get(i, []))

    # Normalize + dedup by cube_no
    seen = set()
    unique = []
    log("Normalize ediliyor...")
    for raw in all_blocks:
        nb = _normalize_block(raw, log)
        if nb is None:
            continue
        cn = nb['cube_no']
        if cn in seen:
            log(f"  tekrar eden cube_no atlandi: {cn}")
            continue
        seen.add(cn)
        unique.append(nb)

    log(f"Toplam {len(unique)} essiz blok cikarildi: {sorted(seen)}")
    # Cache'e kaydet (sonraki ayni PDF'te Gemini cagrilmasin)
    cache_save(pdf_hash, unique, pdf_path)
    log(f"Cache kaydedildi: {CACHE_DIR}\\{pdf_hash}.json")
    return unique


# ---------------- GUI ----------------


class App:
    def __init__(self, root):
        self.root = root
        root.title(f"Last Pages - Cube Block Creator  (v{__version__})")
        root.geometry("900x650")

        top = ttk.Frame(root, padding=10)
        top.pack(fill="x")
        ttk.Label(top, text="Excel:").pack(side="left")
        self.excel_var = tk.StringVar()
        ttk.Entry(top, textvariable=self.excel_var).pack(side="left", fill="x", expand=True, padx=5)
        ttk.Button(top, text="Sec...", command=self.browse).pack(side="left")

        btns = ttk.Frame(root, padding=(10, 0))
        btns.pack(fill="x")
        self.pdf_btn = ttk.Button(btns, text="PDF SEC + GEMINI ILE OKU", command=self.start_pdf_read)
        self.pdf_btn.pack(side="left")
        self.run_btn = ttk.Button(btns, text="EXCEL'E YAZ", command=self.start_run)
        self.run_btn.pack(side="left", padx=5)
        ttk.Button(btns, text="API Key Ayarla", command=self.show_api_key_dialog).pack(side="left", padx=5)
        ttk.Button(btns, text="Cache Yonet", command=self.show_cache_dialog).pack(side="left", padx=5)
        ttk.Button(btns, text="Log temizle", command=self.clear).pack(side="left", padx=5)

        self.info_var = tk.StringVar()
        self._refresh_info()
        info = ttk.Label(root, padding=(10, 5), textvariable=self.info_var, wraplength=880, justify="left")
        info.pack(fill="x")

        self.output = scrolledtext.ScrolledText(root, wrap="word", font=("Consolas", 10))
        self.output.pack(fill="both", expand=True, padx=10, pady=10)

        self.status = ttk.Label(root, text="Hazir.", anchor="w", padding=5)
        self.status.pack(fill="x")

        if os.path.exists(CONFIG):
            try:
                with open(CONFIG, encoding="utf-8") as f:
                    self.excel_var.set(f.read().strip())
            except Exception:
                pass

        self.log(f"Calisma klasoru: {APP_DIR}")
        self.log(f"Python: {sys.version.split()[0]}")
        self.log("")
        self.log("Akis:")
        self.log("  1) 'API Key Ayarla' - Gemini key gir (ilk kez)")
        self.log("  2) 'PDF SEC + GEMINI ILE OKU' - el yazisi formu yukle")
        self.log("  3) 'EXCEL'E YAZ' - Excel'de olmayan cube'lari yazar")
        self.log("")

    def log(self, msg=""):
        self.output.insert("end", msg + "\n")
        self.output.see("end")
        self.root.update_idletasks()

    def clear(self):
        self.output.delete("1.0", "end")

    def browse(self):
        path = filedialog.askopenfilename(
            title="Excel sec",
            filetypes=[("Excel", "*.xlsx *.xls"), ("All", "*.*")],
        )
        if path:
            self.excel_var.set(path)

    def _refresh_info(self):
        key_set = bool(load_api_key())
        key_status = "OK" if key_set else "YOK - 'API Key Ayarla' bas"
        if BLOCKS:
            line1 = f"PDF'ten okunan {len(BLOCKS)} blok: " + ", ".join(str(b['cube_no']) for b in BLOCKS)
        else:
            line1 = "Henuz PDF okunmadi. 'PDF SEC + GEMINI ILE OKU' ile basla."
        self.info_var.set(f"{line1}\nAPI Key: {key_status}   |   Excel'de var olan cube'lar otomatik atlanir.")

    def show_cache_dialog(self):
        win = tk.Toplevel(self.root)
        win.title("Gemini PDF Cache")
        win.geometry("700x400")
        win.transient(self.root)
        win.grab_set()

        ttk.Label(win, padding=10,
                  text=f"Cache klasoru: {CACHE_DIR}\n"
                       "Bir PDF'i daha once okuduysan tekrar Gemini'ye gitmez (para tasarrufu).\n"
                       "Bir entry'i sil -> o PDF'i tekrar yuklersen Gemini'ye tekrar gider.").pack(anchor="w")

        tree = ttk.Treeview(win, columns=("name", "blocks", "date"), show="headings")
        tree.heading("name", text="PDF dosyasi")
        tree.heading("blocks", text="Blok")
        tree.heading("date", text="Tarih")
        tree.column("name", width=350)
        tree.column("blocks", width=60)
        tree.column("date", width=140)
        tree.pack(fill="both", expand=True, padx=10, pady=5)

        entries = []

        def refresh():
            for item in tree.get_children():
                tree.delete(item)
            entries.clear()
            if not os.path.exists(CACHE_DIR):
                return
            for fname in sorted(os.listdir(CACHE_DIR)):
                if not fname.endswith('.json'):
                    continue
                path = os.path.join(CACHE_DIR, fname)
                try:
                    with open(path, encoding='utf-8') as f:
                        d = json.load(f)
                    name = d.get('pdf_name', '(?)')
                    blocks = d.get('block_count', len(d.get('blocks', [])))
                    saved = d.get('saved_at', '?')
                except Exception:
                    name, blocks, saved = fname, '?', '?'
                tree.insert("", "end", values=(name, blocks, saved))
                entries.append(path)

        refresh()

        btn_frame = ttk.Frame(win, padding=10)
        btn_frame.pack(fill="x")

        def delete_selected():
            sel = tree.selection()
            if not sel:
                return
            for item_id in sel:
                idx = tree.index(item_id)
                path = entries[idx]
                try:
                    os.remove(path)
                except Exception:
                    pass
            refresh()
            self.log(f"{len(sel)} cache entry silindi.")

        def delete_all():
            if not messagebox.askyesno("Onay", "Tum cache silinsin mi?", parent=win):
                return
            count = 0
            if os.path.exists(CACHE_DIR):
                for fname in os.listdir(CACHE_DIR):
                    if fname.endswith('.json'):
                        try:
                            os.remove(os.path.join(CACHE_DIR, fname))
                            count += 1
                        except Exception:
                            pass
            refresh()
            self.log(f"Tum cache silindi ({count} entry).")

        ttk.Button(btn_frame, text="Secileni sil", command=delete_selected).pack(side="left")
        ttk.Button(btn_frame, text="Tumunu sil", command=delete_all).pack(side="left", padx=5)
        ttk.Button(btn_frame, text="Kapat", command=win.destroy).pack(side="right")

    def show_api_key_dialog(self):
        win = tk.Toplevel(self.root)
        win.title("Gemini API Key")
        win.geometry("520x180")
        win.transient(self.root)
        win.grab_set()
        ttk.Label(win, text="Gemini API Key:", padding=10).pack(anchor="w")
        var = tk.StringVar(value=load_api_key())
        entry = ttk.Entry(win, textvariable=var, width=70, show="*")
        entry.pack(padx=10, fill="x")
        show_var = tk.BooleanVar(value=False)

        def toggle_show():
            entry.config(show="" if show_var.get() else "*")
        ttk.Checkbutton(win, text="Goster", variable=show_var, command=toggle_show).pack(anchor="w", padx=10, pady=5)
        ttk.Label(win, padding=10, foreground="gray",
                  text="Anahtar yerel olarak gemini_api_key.txt dosyasinda saklanir.\nhttps://aistudio.google.com/apikey adresinden alabilirsin.").pack(anchor="w")

        btn_frame = ttk.Frame(win, padding=10)
        btn_frame.pack(fill="x")

        def do_save():
            key = var.get().strip()
            if not key:
                messagebox.showwarning("Bos", "API key bos olamaz.", parent=win)
                return
            try:
                save_api_key(key)
                self.log(f"API key kaydedildi -> {API_KEY_FILE}")
                self._refresh_info()
                win.destroy()
            except Exception as e:
                messagebox.showerror("Hata", f"Kaydedilemedi: {e}", parent=win)

        ttk.Button(btn_frame, text="Kaydet", command=do_save).pack(side="right")
        ttk.Button(btn_frame, text="Iptal", command=win.destroy).pack(side="right", padx=5)
        entry.focus_set()

    def set_status(self, txt):
        self.status.config(text=txt)
        self.root.update_idletasks()

    def start_pdf_read(self):
        api_key = load_api_key()
        if not api_key:
            messagebox.showwarning("API Key Yok",
                "Once 'API Key Ayarla' butonuyla Gemini API key'i gir.")
            return
        pdf = filedialog.askopenfilename(
            title="PDF sec (el yazisi formu)",
            filetypes=[("PDF", "*.pdf"), ("All", "*.*")],
        )
        if not pdf:
            return
        self.pdf_btn.config(state="disabled")
        self.run_btn.config(state="disabled")
        threading.Thread(target=self._pdf_worker, args=(pdf, api_key), daemon=True).start()

    def _pdf_worker(self, pdf, api_key):
        global BLOCKS
        try:
            self.log("=" * 60)
            self.log(f"PDF okuma basliyor: {pdf}")
            self.set_status("Gemini PDF'i okuyor...")

            def progress(cur, total, phase):
                if phase == "page":
                    self.set_status(f"Gemini sayfa {cur+1}/{total}...")
                elif phase == "render":
                    self.set_status(f"Sayfa render: {cur+1}/{total}")
                elif phase == "gemini":
                    self.set_status("Gemini API cagriliyor...")

            new_blocks = gemini_extract_blocks(pdf, api_key, self.log, progress)
            if not new_blocks:
                raise RuntimeError("Gemini hicbir blok dondurmedi.")

            BLOCKS = new_blocks
            self.log("")
            self.log(f"BLOCKS guncellendi ({len(BLOCKS)} blok). Detay:")
            for b in BLOCKS:
                self.log(f"  cube {b.get('cube_no')}  sample {b.get('sample_mark')}  "
                         f"{len(b.get('rows', []))} satir  batch {b.get('batch_ticket', b.get('batch_tickets'))}")
            self.log("")
            self.log("Simdi 'EXCEL'E YAZ' bas - Excel'de mevcut cube'lar otomatik atlanir.")
            self._refresh_info()
            self.set_status("PDF okundu. EXCEL'E YAZ icin hazir.")
        except Exception as e:
            self.log("")
            self.log(f"*** PDF HATA: {e} ***")
            self.log(traceback.format_exc())
            self.set_status("PDF okuma hatasi.")
            messagebox.showerror("PDF Hata", str(e))
        finally:
            self.pdf_btn.config(state="normal")
            self.run_btn.config(state="normal")

    def start_run(self):
        excel = self.excel_var.get().strip().strip('"')
        if not excel or not os.path.exists(excel):
            messagebox.showerror("Hata", "Excel dosyasi bulunamadi.")
            return
        if not BLOCKS:
            messagebox.showwarning("Blok yok",
                "Once 'PDF SEC + GEMINI ILE OKU' ile PDF yukle.\n"
                "Yazilacak blok yok.")
            return
        if not messagebox.askokcancel("Onay",
                f"{len(BLOCKS)} blok orijinal Excel'in UZERINE yazilacak.\n\n"
                "Yedek aldigindan emin misin?\n\nDosya: " + excel):
            return
        try:
            with open(CONFIG, "w", encoding="utf-8") as f:
                f.write(excel)
        except Exception:
            pass
        self.run_btn.config(state="disabled")
        threading.Thread(target=self._worker, args=(excel,), daemon=True).start()

    def _worker(self, excel):
        try:
            self.log("=" * 60)
            self.set_status("Yaziliyor...")
            run_writer(excel, self.log)
            self.log("")
            self.log("*** BASARILI ***")
            self.set_status("Bitti.")
            messagebox.showinfo("Bitti",
                f"Excel guncellendi:\n{excel}\n\n"
                "Excel'i acip yeni eklenen bloklari (E sutunu kirmizi) kontrol et.")
        except Exception as e:
            self.log("")
            self.log(f"*** HATA: {e} ***")
            self.log(traceback.format_exc())
            self.set_status("Hata.")
            messagebox.showerror("Hata", str(e))
        finally:
            self.run_btn.config(state="normal")


def main():
    root = tk.Tk()
    App(root)
    root.after(800, lambda: check_for_update(root))
    root.mainloop()


if __name__ == "__main__":
    main()
