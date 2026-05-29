"""
create_blocks.py - Cube blocks Excel'e yazma (test surumu).

blocks_data.json'daki blok tanimlarini "Concrete" sheet'ine yazar,
mevcut bloklarin formatini (merge, stil, formul) birebir taklit eder.

Kullanim:
    python create_blocks.py "C:/path/to/Concrete sample form.xlsx"
    python create_blocks.py file.xlsx --start-row 5901
    python create_blocks.py file.xlsx --no-inplace     # ayri dosyaya kaydet
    python create_blocks.py file.xlsx --data blocks_data.json

Varsayilan: orijinal dosyanin UZERINE yazar (yedek aldigindan emin ol).
--start-row verilmezse Concrete sheet'teki son dolu satirin altina yazar.
"""
import argparse
import json
import re
import sys
from copy import copy
from datetime import date, datetime, timedelta
from pathlib import Path

import openpyxl

COL = {
    'A': 1, 'B': 2, 'C': 3, 'D': 4, 'E': 5, 'F': 6, 'G': 7, 'H': 8,
    'I': 9, 'J': 10, 'K': 11, 'L': 12, 'M': 13, 'N': 14, 'O': 15,
    'P': 16, 'Q': 17,
}

WHOLE_BLOCK_MERGES = ['A', 'B', 'E', 'F', 'H', 'J', 'P']
DATE_COLS = ['I', 'K']
DATE_FORMAT = 'DD/MM/YYYY'


def parse_date(s):
    if isinstance(s, datetime):
        return s.date()
    if isinstance(s, date):
        return s
    s = str(s).replace('/', '.').replace('-', '.')
    parts = s.split('.')
    d, m, y = int(parts[0]), int(parts[1]), int(parts[2])
    if y < 100:
        y += 2000
    return date(y, m, d)


def compute_testing_date(sampling, row_type, custom):
    if custom:
        return parse_date(custom)
    days = {
        '7d': 7, 'cmd': 7,
        '28d': 28, 'wp': 28, 'ft': 28,
        '1day': 1, '2day': 2,
    }.get(row_type)
    if days is None:
        return None
    return sampling + timedelta(days=days)


def find_last_data_row(ws):
    for r in range(ws.max_row, 0, -1):
        for col in (1, 2, 3):
            if ws.cell(row=r, column=col).value not in (None, ''):
                return r
    return 0


def find_donor_block(ws, before_row):
    """ Find the cube block ending just before 'before_row'."""
    end = before_row - 1
    while end > 0:
        for col in (1, 2, 3):
            if ws.cell(row=end, column=col).value not in (None, ''):
                break
        else:
            end -= 1
            continue
        break
    if end <= 0:
        return None
    start = end
    for r in range(end, 0, -1):
        if ws.cell(row=r, column=1).value not in (None, ''):
            start = r
            for rng in ws.merged_cells.ranges:
                if rng.min_col == 1 and rng.max_col == 1 and rng.min_row == r:
                    end = max(end, rng.max_row)
                    break
            break
    return (start, end)


def get_donor_styles(ws, start, end):
    n = end - start + 1
    styles = {}
    for col_letter, col_idx in COL.items():
        top = ws.cell(row=start, column=col_idx)
        mid = ws.cell(row=start + n // 2, column=col_idx)
        bot = ws.cell(row=end, column=col_idx)
        styles[col_letter] = {'top': top, 'mid': mid, 'bot': bot}
    return styles


def get_donor_formula(ws, start, end, col_letter):
    col_idx = COL[col_letter]
    for r in range(start, end + 1):
        val = ws.cell(row=r, column=col_idx).value
        if isinstance(val, str) and val.startswith('='):
            return (val, r)
    return (None, None)


def shift_formula(formula, src_row, dst_row):
    diff = dst_row - src_row
    if diff == 0:
        return formula
    def repl(m):
        return f"{m.group(1)}{int(m.group(2)) + diff}"
    return re.sub(r'(\$?[A-Z]+\$?)(\d+)', repl, formula)


def copy_style(src_cell, dst_cell):
    if not src_cell.has_style:
        return
    dst_cell.font = copy(src_cell.font)
    dst_cell.fill = copy(src_cell.fill)
    dst_cell.border = copy(src_cell.border)
    dst_cell.alignment = copy(src_cell.alignment)
    dst_cell.number_format = src_cell.number_format
    dst_cell.protection = copy(src_cell.protection)


def style_position(i, n):
    if i == 0:
        return 'top'
    if i == n - 1:
        return 'bot'
    return 'mid'


def unmerge_overlapping(ws, start_row, end_row):
    to_remove = []
    for rng in list(ws.merged_cells.ranges):
        if rng.min_row <= end_row and rng.max_row >= start_row:
            to_remove.append(str(rng))
    for r in to_remove:
        ws.unmerge_cells(r)


def write_block(ws, block, start_row, donor_styles, donor_formulas):
    rows = block['rows']
    n = len(rows)
    end_row = start_row + n - 1

    unmerge_overlapping(ws, start_row, end_row)
    sampling = parse_date(block['sampling_date'])

    cmd_idx = next((i for i, r in enumerate(rows) if r.get('row_type') == 'cmd'), None)

    batch_tickets = block.get('batch_tickets')
    if not batch_tickets:
        bt = block.get('batch_ticket')
        batch_tickets = [{'ticket': bt, 'rows': [0, n - 1]}] if bt is not None else []

    for i in range(n):
        r = start_row + i
        pos = style_position(i, n)
        for col_letter, col_idx in COL.items():
            copy_style(donor_styles[col_letter][pos], ws.cell(row=r, column=col_idx))
        for dc in DATE_COLS:
            ws.cell(row=r, column=COL[dc]).number_format = DATE_FORMAT

    if n > 1:
        for col_letter in WHOLE_BLOCK_MERGES:
            ws.merge_cells(start_row=start_row, end_row=end_row,
                           start_column=COL[col_letter], end_column=COL[col_letter])

    if cmd_idx is not None and cmd_idx > 0:
        if cmd_idx >= 2:
            ws.merge_cells(start_row=start_row, end_row=start_row + cmd_idx - 1,
                           start_column=COL['D'], end_column=COL['D'])
        if end_row - (start_row + cmd_idx) >= 1:
            ws.merge_cells(start_row=start_row + cmd_idx, end_row=end_row,
                           start_column=COL['D'], end_column=COL['D'])
    elif n > 1:
        ws.merge_cells(start_row=start_row, end_row=end_row,
                       start_column=COL['D'], end_column=COL['D'])

    for bt in batch_tickets:
        a, b = bt['rows']
        if b > a:
            ws.merge_cells(start_row=start_row + a, end_row=start_row + b,
                           start_column=COL['G'], end_column=COL['G'])

    ws.cell(row=start_row, column=COL['A'], value=block['cube_no'])
    ws.cell(row=start_row, column=COL['B'], value=block['sample_mark'])
    ws.cell(row=start_row, column=COL['D'], value=block.get('supplier', 'S2A BP'))
    if cmd_idx is not None and block.get('cmd_code'):
        ws.cell(row=start_row + cmd_idx, column=COL['D'],
                value=f"CMD-{block['cmd_code']}")
    ws.cell(row=start_row, column=COL['E'], value=block.get('site_location', ''))
    if block.get('section') is not None:
        ws.cell(row=start_row, column=COL['F'], value=block['section'])
    for bt in batch_tickets:
        ws.cell(row=start_row + bt['rows'][0], column=COL['G'], value=bt['ticket'])
    ws.cell(row=start_row, column=COL['H'], value=block.get('c_grade', ''))
    ws.cell(row=start_row, column=COL['J'], value=block.get('sampled_by', ''))

    for i, row in enumerate(rows):
        r = start_row + i
        if row.get('mould') is not None:
            ws.cell(row=r, column=COL['C'], value=row['mould'])
        ws.cell(row=r, column=COL['I'], value=sampling)
        td = compute_testing_date(sampling, row.get('row_type'), row.get('testing_date'))
        if td is not None:
            ws.cell(row=r, column=COL['K'], value=td)
        if row.get('age') is not None:
            ws.cell(row=r, column=COL['L'], value=row['age'])
        if row.get('row_type') == 'ft' and row.get('ft_note'):
            ws.cell(row=r, column=COL['N'], value=row['ft_note'])
        if row.get('row_type') == 'site':
            ws.cell(row=r, column=COL['P'], value='Site')
        f, src_r = donor_formulas.get('O', (None, None))
        if f:
            ws.cell(row=r, column=COL['O'], value=shift_formula(f, src_r, r))

    return end_row + 1


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('excel', help='Path to the Excel file')
    ap.add_argument('--start-row', type=int, default=None)
    ap.add_argument('--data', default='blocks_data.json')
    ap.add_argument('--sheet', default='Concrete')
    ap.add_argument('--no-inplace', action='store_true',
                    help='Ayri dosyaya kaydet (varsayilan: orijinal uzerine yaz)')
    args = ap.parse_args()

    data_path = Path(args.data)
    if not data_path.is_absolute():
        data_path = Path(__file__).parent / data_path
    blocks = json.loads(data_path.read_text(encoding='utf-8'))

    wb = openpyxl.load_workbook(args.excel)
    ws = wb[args.sheet]

    start_row = args.start_row if args.start_row else find_last_data_row(ws) + 1
    print(f'Start row: {start_row}')

    donor_range = find_donor_block(ws, start_row)
    if not donor_range:
        sys.exit('ERROR: donor block bulunamadi. Concrete sheet bos mu?')
    donor_start, donor_end = donor_range
    print(f'Donor block: satir {donor_start}-{donor_end}')

    donor_styles = get_donor_styles(ws, donor_start, donor_end)
    donor_formulas = {'O': get_donor_formula(ws, donor_start, donor_end, 'O')}
    if donor_formulas['O'][0]:
        print(f"Donor O formulu: {donor_formulas['O'][0]} (satir {donor_formulas['O'][1]})")

    next_row = start_row
    for block in blocks:
        print(f"  -> cube {block['cube_no']} @ satir {next_row} ({len(block['rows'])} satir)")
        next_row = write_block(ws, block, next_row, donor_styles, donor_formulas)

    if args.no_inplace:
        p = Path(args.excel)
        out = str(p.with_name(p.stem + '.with_blocks' + p.suffix))
    else:
        out = args.excel
    wb.save(out)
    print(f'\nKaydedildi: {out}')


if __name__ == '__main__':
    main()
